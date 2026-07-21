import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def make_excel():
    json_path = r"c:\Users\서승재\OneDrive\Desktop\네이버 자동화\6월22일\고속도로\scratch\excel-data.json"
    if not os.path.exists(json_path):
        print("JSON file not found.")
        return
        
    with open(json_path, "r", encoding="utf-8") as f:
        rows_data = json.load(f)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고속도로 전체 페이지 목록"
    
    # Styling
    header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers = ["제목", "주소", "CTA문구 링크"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    font_data = Font(name="Malgun Gothic", size=10)
    
    # Write data rows
    row_num = 2
    for r in rows_data:
        # A: 제목
        ws.cell(row=row_num, column=1, value=r["title"]).font = font_data
        ws.cell(row=row_num, column=1).alignment = left_align
        
        # B: 주소
        ws.cell(row=row_num, column=2, value=r["url"]).font = font_data
        ws.cell(row=row_num, column=2).alignment = left_align
        
        # C: CTA문구 링크
        ws.cell(row=row_num, column=3, value=r["cta"]).font = font_data
        ws.cell(row=row_num, column=3).alignment = left_align
        
        row_num += 1
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            val_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 100) # Cap at 100 to avoid overly wide columns
        
    output_path = r"C:\Users\서승재\Downloads\고속도로4_최종3.xlsx"
    try:
        wb.save(output_path)
        print(f"Successfully generated final Excel file at: {output_path}")
        print(f"Total rows written: {len(rows_data)}")
    except PermissionError:
        fallback_path = r"C:\Users\서승재\Downloads\고속도로4_최종3_new.xlsx"
        try:
            wb.save(fallback_path)
            print(f"Successfully generated final Excel file (fallback) at: {fallback_path}")
            print(f"Total rows written: {len(rows_data)}")
        except Exception as e2:
            print(f"Error saving fallback Excel file: {e2}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

if __name__ == "__main__":
    make_excel()
