import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def generate_excel():
    sitemap_path = r"c:\Users\서승재\OneDrive\Desktop\네이버 자동화\6월22일\고속도로\out\sitemap.xml"
    if not os.path.exists(sitemap_path):
        # Fallback to public/sitemap.xml
        sitemap_path = r"c:\Users\서승재\OneDrive\Desktop\네이버 자동화\6월22일\고속도로\public\sitemap.xml"
        
    if not os.path.exists(sitemap_path):
        print("Sitemap file not found.")
        return

    print(f"Reading sitemap from: {sitemap_path}")
    with open(sitemap_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Extract all <loc> contents
    urls = re.findall(r"<loc>(.*?)</loc>", content)
    print(f"Total URLs found: {len(urls)}")

    # Initialize workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고속도로 포털 페이지 목록"

    # Header styling
    header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Set headers
    headers = ["번호", "구분 (카테고리)", "페이지 주소 (URL)", "비고"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Helper function to categorize URLs
    def get_category(url):
        if url.endswith("/traffic") or url.endswith("/traffic/"):
            return "교통상황 인덱스"
        elif "/traffic/" in url:
            return "실시간 교통상황 (고속도로 노선별)"
        elif url.endswith("/gas") or url.endswith("/gas/"):
            return "주유소 인덱스"
        elif "/gas/" in url:
            return "실시간 주유소 가격비교 (휴게소별)"
        elif url.endswith("/ev") or url.endswith("/ev/"):
            return "전기차 인덱스"
        elif "/ev/" in url:
            return "전기차 충전소 가격비교 (휴게소별)"
        elif "/rest-areas/" in url:
            if url.endswith("/food") or url.endswith("/food/"):
                return "휴게소 실시간 맛집메뉴"
            elif url.endswith("/facilities") or url.endswith("/facilities/"):
                return "휴게소 편의시설 정보"
            else:
                return "휴게소 상세 정보 대시보드"
        elif "/region/" in url:
            return "지역별 휴게소 목록"
        elif url.endswith(".kr") or url.endswith(".kr/"):
            return "메인 홈페이지"
        else:
            return "기타 정보 페이지"

    font_data = Font(name="Malgun Gothic", size=10)
    
    # Write data rows
    row_num = 2
    category_counts = {}
    for idx, url in enumerate(urls, 1):
        category = get_category(url)
        category_counts[category] = category_counts.get(category, 0) + 1
        
        # Row data
        ws.cell(row=row_num, column=1, value=idx).font = font_data
        ws.cell(row=row_num, column=1).alignment = center_align
        
        ws.cell(row=row_num, column=2, value=category).font = font_data
        ws.cell(row=row_num, column=2).alignment = center_align
        
        ws.cell(row=row_num, column=3, value=url).font = font_data
        ws.cell(row=row_num, column=3).alignment = left_align
        
        ws.cell(row=row_num, column=4, value="").font = font_data
        
        row_num += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            # Korean characters count as 2 length
            val_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save to Downloads folder
    output_path = r"C:\Users\서승재\Downloads\고속도로_전체_페이지.xlsx"
    try:
        wb.save(output_path)
        print(f"Successfully generated Excel file at: {output_path}")
        print("\nCategory counts:")
        for cat, cnt in category_counts.items():
            print(f"- {cat}: {cnt}")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    generate_excel()
